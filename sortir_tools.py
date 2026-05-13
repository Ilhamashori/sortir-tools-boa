# ============================================================
#  BOA SORTIR TOOLS  —  Standalone
#  Copyright (c) 2026 Ilham Mashori — Beauty of Angel
#  All Rights Reserved.
# ============================================================
#
#  Upload PDF resi acak → output ZIP (PDF per kurir, tersortir SKU+qty)
#  + Fitur Kloteran: auto-assign batch per produk/kurir
#  Tidak bergantung app.py — standalone dengan login sendiri.
# ============================================================

import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
import json
import zipfile
from datetime import datetime, timedelta
from collections import defaultdict
import math

# ══════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Sortir Tools - BOA",
    page_icon="🗂️",
    layout="centered",
)

st.markdown("""
<style>
#MainMenu {visibility: hidden;}
footer    {visibility: hidden;}
header    {visibility: hidden;}
[data-testid="stToolbar"] {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
#  GLOBAL THEME — Warm Natural Dark
# ══════════════════════════════════════════════════════════════

st.markdown("""
<style>
.stApp { background: #15110D !important; color: #F0E4D2 !important; }
.stApp > div { background: #15110D !important; }
h1, h2, h3, h4, h5, h6 { color: #F0E4D2 !important; }
p, span, label, div { color: #E8DCC9; }
section[data-testid="stSidebar"] { background: #1A1411 !important; }
div[data-testid="stButton"] button {
    background: #C49166 !important; color: #15110D !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; transition: background 0.15s ease;
}
div[data-testid="stButton"] button:hover { background: #D4A574 !important; }
div[data-testid="stButton"] button[kind="secondary"] {
    background: #2E251E !important; color: #F0E4D2 !important;
    border: 0.5px solid #4A3A2C !important;
}
div[data-testid="stTextInput"] label,
div[data-testid="stFileUploader"] label,
.stCheckbox label p { color: #C9B89F !important; font-weight: 500 !important; }
div[data-testid="stTextInput"] input {
    background: #1F1815 !important; border: 0.5px solid #3A2F26 !important;
    border-radius: 8px !important; color: #F0E4D2 !important; font-size: 14px !important;
}
div[data-testid="stTextInput"] input::placeholder { color: #6B5D4D !important; }
div[data-testid="stTextInput"] input:focus {
    border-color: #C49166 !important;
    box-shadow: 0 0 0 2px rgba(196,145,102,0.25) !important;
}
[data-testid="stFileUploader"] section {
    background: #1F1815 !important; border: 1px dashed #3A2F26 !important;
    border-radius: 10px !important;
}
[data-testid="stFileUploader"] section button { background: #C49166 !important; color: #15110D !important; }
div[data-testid="stDownloadButton"] button {
    background: #C49166 !important; color: #15110D !important;
    border-radius: 8px !important; font-weight: 600 !important;
}
div[data-testid="stDownloadButton"] button:hover { background: #D4A574 !important; }
.stProgress > div > div > div > div { background-color: #C49166 !important; }
.stProgress > div > div > div { background-color: #2E251E !important; }
div[data-testid="stAlert"] {
    background: #2A1A12 !important; border: 0.5px solid #4A2818 !important;
    color: #E8B89F !important; border-radius: 8px !important;
}
div[data-baseweb="notification"] { background: #2A1A12 !important; color: #E8B89F !important; }
.stDataFrame { background: #1A1411 !important; border-radius: 8px; }
hr { border-color: #3A2F26 !important; }
[data-testid="stCaptionContainer"], small { color: #8A7A66 !important; }
.stCheckbox [role="checkbox"][aria-checked="true"] {
    background-color: #C49166 !important; border-color: #C49166 !important;
}
.streamlit-expanderHeader {
    background: #1F1815 !important; color: #F0E4D2 !important; border-radius: 8px !important;
}
.stSpinner > div { border-top-color: #C49166 !important; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
#  LOAD CONFIG
# ══════════════════════════════════════════════════════════════

@st.cache_data
def load_config():
    try:
        with open("rules_sortir.json") as f:
            return json.load(f)
    except FileNotFoundError:
        return {"users": {"sortirboa": {"password": "sortir2026"}}}

CONFIG = load_config()
USERS  = CONFIG.get("users", {})

# ══════════════════════════════════════════════════════════════
#  LOGIN
# ══════════════════════════════════════════════════════════════

def _do_login(username: str, password: str) -> bool:
    u = username.strip().lower()
    if u in USERS and USERS[u]["password"] == password:
        st.session_state["logged_in"] = True
        st.session_state["username"]  = u
        return True
    return False

def _logout():
    for k in ["logged_in", "username"]:
        st.session_state.pop(k, None)

if not st.session_state.get("logged_in"):
    st.markdown("""
    <style>
        @media (max-width: 640px) {
            div[data-testid="column"] { width: 100% !important; flex: 1 1 100% !important; }
        }
    </style>
    """, unsafe_allow_html=True)

    col_brand, col_form = st.columns([1, 1], gap="medium")

    with col_brand:
        st.markdown("""
        <div style="background:#1F1815;padding:1.75rem 1.5rem;color:#F0E4D2;
                    border-radius:14px;min-height:360px;display:flex;
                    flex-direction:column;justify-content:space-between;
                    border:0.5px solid #2E251E;
                    box-shadow:0 4px 20px rgba(0,0,0,0.25);">
            <div>
                <div style="display:flex;align-items:center;gap:10px;margin-bottom:1.25rem;">
                    <div style="width:36px;height:36px;border-radius:10px;
                                background:linear-gradient(135deg,#C49166 0%,#A8744E 100%);
                                display:flex;align-items:center;justify-content:center;
                                color:#15110D;font-weight:700;font-size:18px;
                                box-shadow:0 2px 8px rgba(196,145,102,0.3);">S</div>
                    <div>
                        <div style="font-weight:600;font-size:14px;color:#F0E4D2;">BOA Sortir Tools</div>
                        <div style="font-size:11px;color:#D4A574;">Pemisah Resi per Ekspedisi</div>
                    </div>
                </div>
                <div style="font-size:22px;font-weight:600;color:#F0E4D2;line-height:1.4;margin-bottom:.5rem;">
                    Sortir resi otomatis,<br/>1 klik, 1 ZIP.
                </div>
                <div style="font-size:12px;color:#B8A78D;line-height:1.6;">
                    Upload 1 PDF resi campuran dari Kiriminaja → otomatis dipisah per kurir
                    (SPX, SiCepat, J&T, JNE, dst.) → download ZIP berisi PDF per kurir.
                </div>
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:1.25rem;">
                <div style="background:#15110D;border:0.5px solid #2E251E;border-radius:8px;padding:10px 12px;">
                    <div style="font-size:10px;color:#D4A574;text-transform:uppercase;letter-spacing:.5px;">Versi</div>
                    <div style="font-size:16px;font-weight:600;color:#F0E4D2;margin-top:2px;">v1.1</div>
                </div>
                <div style="background:#15110D;border:0.5px solid #2E251E;border-radius:8px;padding:10px 12px;">
                    <div style="font-size:10px;color:#D4A574;text-transform:uppercase;letter-spacing:.5px;">Status</div>
                    <div style="font-size:16px;font-weight:600;color:#F0E4D2;margin-top:2px;">Production</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_form:
        st.markdown("""
        <div style="display:inline-flex;align-items:center;background:#2A1F18;
                    color:#D4A574;font-size:10px;font-weight:600;padding:4px 10px;
                    border-radius:999px;text-transform:uppercase;letter-spacing:.5px;
                    margin-bottom:10px;border:0.5px solid #4A3A2C;">Akses Internal</div>
        <div style="font-size:19px;font-weight:600;color:#F0E4D2;margin:0 0 4px 0;">
            Selamat datang
        </div>
        <div style="font-size:12px;color:#8A7A66;margin:0 0 1rem 0;">
            Login untuk akses tools sortir ekspedisi
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form", clear_on_submit=False):
            uname = st.text_input("Username", placeholder="Username", key="login_user")
            pw    = st.text_input("Password", type="password", placeholder="Password", key="login_pass")
            submitted = st.form_submit_button("Masuk", use_container_width=True, type="primary")
        if submitted:
            if _do_login(uname, pw):
                st.rerun()
            else:
                st.error("❌ Username atau password salah.")

        st.markdown("""
        <div style="display:flex;justify-content:space-between;
                    font-size:11px;color:#6B5D4D;margin-top:14px;">
            <span>Beauty of Angel © 2026</span>
            <span>Hubungi admin untuk akses</span>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("""
    <div style="margin-top:14px;text-align:center;font-size:10px;color:#6B5D4D;
                letter-spacing:0.2px;">
        Crafted by
        <span onclick="
            var p = document.getElementById('creditPopover');
            p.style.display = (p.style.display === 'block') ? 'none' : 'block';
        " style="color:#D4A574;font-weight:500;cursor:pointer;
                 border-bottom:1px dotted #3A2F26;padding-bottom:1px;">Mashori</span>
        <div id="creditPopover" style="display:none;margin:8px auto 0;
             padding:10px 14px;background:#1F1815;border:0.5px solid #2E251E;
             border-radius:10px;font-size:11px;color:#B8A78D;max-width:320px;">
            <div style="margin-bottom:6px;color:#F0E4D2;">
                Butuh tools custom untuk bisnis Anda?
            </div>
            <div style="display:flex;gap:14px;justify-content:center;">
                <a href="https://wa.me/6281296547033" target="_blank"
                   style="color:#C49166;text-decoration:none;font-weight:500;">WhatsApp</a>
                <span style="color:#3A2F26;">·</span>
                <a href="mailto:ilhammashori@gmail.com"
                   style="color:#C49166;text-decoration:none;font-weight:500;">Email</a>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.stop()

# ══════════════════════════════════════════════════════════════
#  KURIR DETECTION RULES
# ══════════════════════════════════════════════════════════════

KURIR_RULES = [
    {"kurir": "SPX", "patterns": [r"\bspxid\d{6,}\b", r"shopee\s*express", r"\bspx\b"], "priority": 10},
    {"kurir": "SiCepat", "patterns": [
        r"\bsicepat\b", r"\bsi\s*cepat\b", r"\bsicep[a-z/]{0,3}t?\b",
        r"\bhalu\s*5\b", r"\bhalu\s*ribu\b", r"halu.{0,3}5.{0,3}ribu",
        r"\bribu\b", r"\b0046\d{8}\b"], "priority": 9},
    {"kurir": "JNT LEX", "patterns": [r"\bjnt\s*lex\b", r"\bJNTLEX\d+\b"], "priority": 10},
    {"kurir": "J&T", "patterns": [
        r"\bj\s*&\s*t\b", r"\bjnt\s*express\b", r"\bjnt\b",
        r"\bjet\.co\.id\b", r"\b(?:JX|JT|JP)\d{8,}\b"], "priority": 9},
    {"kurir": "JNE LEX", "patterns": [r"\bjne\s*lex\b", r"\bjnex\b"], "priority": 10},
    {"kurir": "JNE", "patterns": [r"\bjne\s*(?:express|reg|yes|oke|trucking)?\b"], "priority": 9},
    {"kurir": "Anteraja", "patterns": [
        r"\banteraja\b", r"\bantar\s*aja\b", r"pakeko\s*aja",
        r"\bpakeko\b", r"\bPLBX?\d*\b", r"\bTSA-\d{6,}\b"], "priority": 9},
    {"kurir": "GTL", "patterns": [r"\bgtl\b", r"goto\s*logistics"], "priority": 9},
    {"kurir": "Ninja Xpress", "patterns": [r"\bninja\s*xpress\b", r"\bninjaxpress\b", r"\blnid\d+\b"], "priority": 9},
    {"kurir": "Ninja LEX", "patterns": [r"\bninja\s*lex\b", r"\bnlex\b"], "priority": 10},
    {"kurir": "SAP", "patterns": [r"\bsap\s*express\b", r"\bsap\b"], "priority": 8},
    {"kurir": "SAP LEX", "patterns": [r"\bsap\s*lex\b"], "priority": 10},
    {"kurir": "Ninja Pusat", "patterns": [r"\bninja\s*pusat\b", r"\bninjapusat\b"], "priority": 10},
    {"kurir": "ID Express", "patterns": [r"\bid\s*express\b", r"\bidexpress\b", r"\bidexp\b"], "priority": 9},
    {"kurir": "Pos Indonesia", "patterns": [r"\bpos\s*indonesia\b", r"\bpt\s*pos\b", r"\bpos\b"], "priority": 8},
    {"kurir": "Lion Parcel", "patterns": [r"\blion\s*parcel\b", r"\blionparcel\b", r"\blion\b"], "priority": 8},
    {"kurir": "Instan", "patterns": [r"\binstan\b", r"\bsameday\b", r"\bsame\s*day\b", r"\binstant\b"], "priority": 9},
    {"kurir": "Sentral Cargo", "patterns": [r"\bsentral\s*cargo\b", r"\bsentral\b"], "priority": 9},
]

KURIR_LIST             = [r["kurir"] for r in KURIR_RULES]
KURIR_BELUM_TERDETEKSI = "Belum-Terdeteksi"


def detect_kurir(text: str) -> str | None:
    if not text:
        return None
    for rule in sorted(KURIR_RULES, key=lambda r: -r["priority"]):
        for pat in rule["patterns"]:
            if re.search(pat, text, re.IGNORECASE):
                return rule["kurir"]
    return None


# ══════════════════════════════════════════════════════════════
#  OCR FALLBACK
# ══════════════════════════════════════════════════════════════

def _render_page_png(pdf_bytes: bytes, page_index: int, dpi: int = 200) -> bytes | None:
    try:
        import fitz
        doc  = fitz.open(stream=pdf_bytes, filetype="pdf")
        page = doc[page_index]
        pix  = page.get_pixmap(dpi=dpi)
        out  = pix.tobytes("png")
        doc.close()
        return out
    except Exception:
        return None


def _ocr_logo(png_bytes: bytes) -> str:
    try:
        import pytesseract
        from PIL import Image, ImageEnhance
        img  = Image.open(io.BytesIO(png_bytes))
        w, h = img.size
        crops = [
            img.crop((int(w * .68), 0, w, int(h * .07))),
            img.crop((int(w * .50), 0, w, int(h * .10))),
        ]
        results = []
        for c in crops:
            for v in [c.convert("L"), ImageEnhance.Contrast(c.convert("L")).enhance(2.5)]:
                for psm in (6, 7, 11):
                    try:
                        t = pytesseract.image_to_string(v, config=f"--psm {psm}")
                        if t.strip():
                            results.append(t)
                    except Exception:
                        pass
        return " ".join(results)
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════
#  GROQ VISION FALLBACK (Layer 3)
# ══════════════════════════════════════════════════════════════

def _get_groq_api_key() -> str | None:
    try:
        return st.secrets.get("GROQ_API_KEY", None)
    except Exception:
        return None


def _groq_detect_kurir(png_bytes: bytes, api_key: str | None = None) -> str | None:
    if not api_key:
        api_key = _get_groq_api_key()
    if not api_key:
        return None
    try:
        from groq import Groq
        import base64
        b64 = base64.b64encode(png_bytes).decode()
        client = Groq(api_key=api_key)
        kurir_options = [r["kurir"] for r in KURIR_RULES]
        prompt = (
            "Lihat logo ekspedisi/kurir di label pengiriman ini (biasanya pojok kanan-atas). "
            f"Identifikasi 1 kurir dari pilihan: {', '.join(kurir_options)}. "
            "Catatan: logo HALU = SiCepat (HALU adalah sub-brand SiCepat). "
            "Jawab HANYA nama kurirnya saja (1 kata/frasa singkat), tanpa penjelasan. "
            "Kalau tidak yakin atau logo tidak jelas, jawab 'Unknown'."
        )
        resp = client.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[{"role": "user", "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
            ]}],
            max_tokens=20, temperature=0,
        )
        result = (resp.choices[0].message.content or "").strip()
        result_lower = result.lower()
        if "unknown" in result_lower or not result_lower:
            return None
        for k in kurir_options:
            kl = k.lower()
            if kl in result_lower or kl.replace(" ", "") in result_lower.replace(" ", ""):
                return k
        if "halu" in result_lower or "sicep" in result_lower: return "SiCepat"
        if "jnt" in result_lower or "j&t" in result_lower: return "J&T"
        if "shopee" in result_lower: return "SPX"
        if "ninja" in result_lower: return "Ninja Xpress"
        return None
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════
#  PARSE PRODUK
# ══════════════════════════════════════════════════════════════

# ── SKU / keyword produk yg harus diabaikan (free gift, packing) ──────────────
# Kalau SKU utama dari merchant table masuk list ini → parse_produk return {}
# Kalau teks label mengandung baris free-gift → baris itu di-strip dulu
_IGNORE_SKU_EXACT: set[str] = {
    "BUBBLE01",
    "0101ASRT",                                   # amplop lebaran
    "0505PHAROS", "0505MAGIA", "0505ABROWIE",
    "0505SPEXSYMBOL", "0505LAUSHINE", "0505LANIER",
    "0505Brow&Lash",
}

# Regex prefix 0505 (catch kolaborasi baru yg belum ada di list di atas)
_IGNORE_SKU_PREFIX_RE = re.compile(r"^0505", re.IGNORECASE)

# Keyword di JUDUL BARIS yg menandakan baris itu adalah free-gift → dibuang dari teks
# sebelum PRODUK_PATTERNS di-scan
_IGNORE_LINE_RE = re.compile(
    r"product\s+grtis|free\s+(?:gift|rndom)|"
    r"bubble\s+wrap|bubble\s+packing|plastik.{0,10}packing|"
    r"amplop\s+lebaran|colabora(?:si|tion)|exclusive\s+gift",
    re.IGNORECASE,
)

# ── Text patterns → nama produk (urutan penting: spesifik dulu) ──────────────
PRODUK_PATTERNS = [
    # Soap / Lotion Sachet — deteksi sebelum kata generik
    (r"glow\s*soap",                    "Glow Soap"),
    (r"body\s*lotion\s*sachet",         "Body Lotion Sachet"),

    # Serum spesifik — sebelum generic 'serum'
    (r"brightening\s*serum",            "Brightening Serum"),
    (r"brightening\s*advance",          "Brightening Serum"),   # varian judul lain
    (r"acne\s*serum",                   "Acne Serum"),
    (r"peeling\s*(?:serum|advance|solution)", "Peeling Serum"),
    (r"moisturi[sz]er\s*serum",         "Moisturizer Serum"),   # fix: pola lama moist\s*serum tdk cocok
    (r"moisturizer",                    "Moisturizer Serum"),   # fallback jika tanpa kata serum

    # Body care
    (r"body\s*lotion",                  "Body Lotion"),
    (r"body\s*wash",                    "Body Wash"),

    # Lip & Eye
    (r"lip\s*serum",                    "Lip Serum"),
    (r"eye\s*cre?am",                   "Eye Cream"),
    (r"eye\s*wrin[kl]",                 "Eye Wrinkle"),         # Eye Wrinkel / Eye Wrinkle

    # Underarm
    (r"underarm",                       "Underarm"),

    # Feminine (spray & nature)
    (r"feminin",                        "Feminine"),            # cocok: Feminine Spray, Nature Feminine

    # Sunscreen
    (r"sunscreen",                      "Sunscreen"),

    # White Tomato Series — sebelum 'facial wash' / 'day cream' dst agar tidak pecah
    (r"white\s*tomato",                 "White Tomato Series"),
    (r"\bwt\s+(?:series|brightening|sunscreen|day|night|facial|toner)\b", "White Tomato Series"),

    # Parfume BYC (by Umar Wijaya / kolaborasi)
    (r"\bbyc\b",                        "Parfume BYC"),

    # Parfume BOA (extrait / individual variant)
    (r"extrait\s*de\s*parfum",          "Parfume"),
    (r"parfum[e]?",                     "Parfume"),

    # WT sub-products (hanya jika WT Series belum ke-match)
    (r"face\s*toner|brightening\s*toner", "Toner"),
    (r"facial\s*wash",                  "Facial Wash"),
    (r"day\s*cream",                    "Day Cream"),
    (r"night\s*gel",                    "Night Gel"),

    # Collagen Drink — sebelum 'drink' saja
    (r"collagen\s*drink",               "Collagen Drink"),
    (r"collagen\s+(?:gluthatione|glutathione)", "Collagen Drink"),  # judul tanpa kata 'drink'

    # Cleansing Balm
    (r"cleansing\s*balm",               "Cleansing Balm"),

    # Catch-all BOA (dihapus di post-processing, tapi berguna sbg fallback)
    (r"beauty\s*of\s*angel",            "BOA Produk"),
]

# ── SKU → nama produk (dibangun dari data BigSeller nyata) ───────────────────
# Urutan penting: pola lebih spesifik/panjang DULU sebelum pola generik
_SKU_TO_PRODUK = [
    # ── Lip Serum (0310) ────────────────────────────────────────────────────
    (r"^0310[A-Z]?MLP",                 "Lip Serum"),       # 0310AMLP, 0310BMLP, dst
    (r"^0310MLPEWC",                    "Lip Serum"),       # bundle LP + Eye Wrinkle

    # ── Glow Soap (0104) ────────────────────────────────────────────────────
    (r"^0104[A-Z0-9-]*GSR",             "Glow Soap"),       # semua varian GSR
    (r"^0104SPONS",                     "Glow Soap"),       # 4 PCS Lebih Hemat
    (r"^0207ABWPC$",                    "Glow Soap"),       # listing glow soap dgn prefix 0207

    # ── Brightening Serum (0105) ─────────────────────────────────────────────
    (r"^0105[A-Z]+BTS",                 "Brightening Serum"),

    # ── Peeling Serum / Solution (0105) ──────────────────────────────────────
    (r"^0105[A-Z]+APS|^0105[A-Z]+PS\b", "Peeling Serum"),

    # ── Moisturizer Serum (0105) ──────────────────────────────────────────────
    (r"^0105[A-Z]+MOIS",                "Moisturizer Serum"),

    # ── Acne Serum (0105) ────────────────────────────────────────────────────
    (r"^0105[A-Z]+CCS",                 "Acne Serum"),

    # ── Eye Cream (0202 …EYC) ────────────────────────────────────────────────
    (r"^0202[A-Z]+EYC",                 "Eye Cream"),

    # ── Eye Wrinkle (0202 …EWC) ──────────────────────────────────────────────
    (r"^0202[A-Z]+EWC",                 "Eye Wrinkle"),

    # ── Body Lotion Sachet — HARUS sebelum Body Lotion ──────────────────────
    (r"^0208[A-Z]+BLTS",                "Body Lotion Sachet"),
    (r"^0208FABLES",                    "Body Lotion Sachet"),  # 1pcs lotion + 5 sachet

    # ── Body Lotion (0208 …BLT) ──────────────────────────────────────────────
    (r"^0208[A-Z]+BLT",                 "Body Lotion"),

    # ── Body Wash (0207) ─────────────────────────────────────────────────────
    (r"^0207[A-Z0-9]+BWP",              "Body Wash"),
    (r"^0207[A-Z0-9]+BWC",              "Body Wash"),
    (r"^0207BBWCDS",                    "Body Wash"),

    # ── Underarm (0208 …AU / BUC / CUC / BUS) ───────────────────────────────
    (r"^0208[A-Z0-9-]*(AUC|AUCN|AUS|AUSN|BUCNS?A?|CUCN|BUCNG|BUSN)",
                                        "Underarm"),

    # ── Sunscreen ────────────────────────────────────────────────────────────
    (r"^0101[A-Z]SSC",                  "Sunscreen"),
    (r"^0101[A-Z]SPF",                  "Sunscreen"),
    (r"^0203[A-Z]SPF",                  "Sunscreen"),          # paket WT sunscreen

    # ── White Tomato Series ──────────────────────────────────────────────────
    (r"^0100[A-Z]+WT",                  "White Tomato Series"),
    (r"^0101[A-Z]+WT",                  "White Tomato Series"),
    (r"^0104[A-Z]+WT",                  "White Tomato Series"),
    (r"^0203[A-Z]+WT",                  "White Tomato Series"),

    # ── Parfume BYC ──────────────────────────────────────────────────────────
    (r"^0209[A-Z0-9-]*BYC",             "Parfume BYC"),
    (r"^0209EBYC",                      "Parfume BYC"),

    # ── Parfume BOA (individual & bundle) ────────────────────────────────────
    (r"^0209[A-Z]+P(BT|DM|FL|EP)",      "Parfume"),   # single: Butter/Dynamite/FirstLove/Euphoria
    (r"^0209[A-Z]+PDM",                 "Parfume"),   # 0209BPDM-1 dst
    (r"^0209[A-Z]+BB(TDM|TEP|TFL)",     "Parfume"),   # bundle 2 parfume BOA
    (r"^0209[A-Z]+BD(MEP|MFL)",         "Parfume"),   # bundle Dynamite + …
    (r"^0209[A-Z]+BFL(EP)?",            "Parfume"),
    (r"^0209[A-Z0-9]+(CPB|CPD|CPF|DPB|DPD|DPF|DPE|BPB|BPD|BPF)", "Parfume"),
    (r"^0290",                          "Parfume"),   # 0290BPEP, 0290CPEP, 0290DPEP

    # ── Feminine ─────────────────────────────────────────────────────────────
    (r"^0209[A-Z]+NFN",                 "Feminine"),
    (r"^0209[A-Z]+FSP",                 "Feminine"),

    # ── Collagen Drink (0412) ────────────────────────────────────────────────
    (r"^0412[A-Z0-9-]+CDM",             "Collagen Drink"),

    # ── Cleansing Balm ───────────────────────────────────────────────────────
    (r"^0106ACSB",                      "Cleansing Balm"),

    # ── Men Care ─────────────────────────────────────────────────────────────
    (r"^0213AMCR",                      "Men Care"),
    (r"^MENCARE$",                       "Men Care"),

    # ── Toner ────────────────────────────────────────────────────────────────
    (r"^0203[A-Z]+TNR",                 "Toner"),
]


def _is_ignore_sku(sku_code: str) -> bool:
    """Return True jika SKU adalah free-gift / packing yg harus diabaikan."""
    sku_up = sku_code.strip().upper()
    if sku_up in {s.upper() for s in _IGNORE_SKU_EXACT}:
        return True
    if _IGNORE_SKU_PREFIX_RE.match(sku_code.strip()):
        return True
    return False


def _sku_to_produk_name(sku_code: str) -> str | None:
    """Return nama produk dari SKU, atau None kalau tidak dikenal / harus diabaikan."""
    if _is_ignore_sku(sku_code):
        return None
    for pattern, nama in _SKU_TO_PRODUK:
        if re.search(pattern, sku_code, re.IGNORECASE):
            return nama
    return None


def _strip_ignore_lines(text: str) -> str:
    """Buang baris yg mengandung keyword free-gift agar tidak ikut dideteksi sebagai produk."""
    lines = text.split("\n")
    cleaned = [ln for ln in lines if not _IGNORE_LINE_RE.search(ln)]
    return "\n".join(cleaned)


def _parse_merchant_table(text_clean: str) -> tuple:
    merch_m = re.search(
        r"Merchant\s+Title\s+SKU\s+Qty(.+)",
        text_clean, re.DOTALL | re.IGNORECASE
    )
    if not merch_m:
        return None, None, None
    section = merch_m.group(1)
    first_companion = None   # fallback: companion jadi primary kalau tidak ada produk lain

    for line in section.split("\n"):
        line = line.strip()
        if not line or re.search(r"qty\s+total", line, re.IGNORECASE):
            continue
        # Skip baris free-gift dalam merchant table
        if _IGNORE_LINE_RE.search(line):
            continue
        m = re.search(r"^(.+?)\s+\b([A-Z0-9]{6,}[-A-Z0-9]*)\s+(\d+)\s*$", line)
        if m:
            sku = m.group(2).strip()
            if _is_ignore_sku(sku):
                continue
            # Companion SKU → skip untuk primary, simpan sebagai fallback
            if sku.upper() in {k.upper() for k in _COMPANION_SKUS}:
                if first_companion is None:
                    first_companion = (m.group(1).strip(), sku, int(m.group(3)))
                continue
            return m.group(1).strip(), sku, int(m.group(3))

    # Tidak ada produk utama → pakai companion sebagai primary (e.g. resi Men Care murni)
    if first_companion:
        return first_companion
    return None, None, None


def _base_produk(sku_name: str) -> str:
    return re.sub(r"\s*\d+\s*PCS\s*$", "", sku_name, flags=re.IGNORECASE).strip()


def parse_produk(text: str) -> dict:
    if not text:
        return {}

    text_clean = re.sub(r"(?:Pengirim|Sender)\s*[:\(][^\n]*\n?", "", text, flags=re.IGNORECASE)
    merch_title, table_sku, table_qty = _parse_merchant_table(text_clean)

    if table_sku:
        sku_produk = _sku_to_produk_name(table_sku)
        if sku_produk:
            qty = table_qty if table_qty else 1
            key = f"{sku_produk} {qty} PCS" if qty > 1 else sku_produk
            return {key: 1}

    # Buang baris free-gift sebelum text-scan
    search_text = _strip_ignore_lines(text_clean)
    if merch_title:
        search_text = search_text + "\n" + merch_title

    barang_m = re.search(r"Barang\s*:\s*(.+?)(?:\n|$)", text_clean, re.IGNORECASE)

    hasil        = {}
    matched_base: set[str] = set()

    for pat, nama in PRODUK_PATTERNS:
        if not re.search(pat, search_text, re.IGNORECASE):
            continue
        nama_upper = nama.upper()
        if any(matched.startswith(nama_upper + " ") for matched in matched_base):
            continue
        matched_base.add(nama_upper)

        if table_qty is not None:
            qty = table_qty
        elif barang_m:
            m_qty = re.search(r"(\d+)\s*PC[SO]+", barang_m.group(1), re.IGNORECASE)
            qty = int(m_qty.group(1)) if m_qty else 1
        else:
            all_matches = list(re.finditer(pat, search_text, re.IGNORECASE))
            last_m = all_matches[-1]
            window = search_text[max(0, last_m.start() - 40): last_m.end() + 60]
            m_qty = re.search(r"(\d+)\s*PC[SO]+", window, re.IGNORECASE)
            qty = int(m_qty.group(1)) if m_qty else 1

        key = f"{nama} {qty} PCS" if qty > 1 else nama
        hasil[key] = hasil.get(key, 0) + 1

    if len(hasil) > 1:
        bl_keys = [k for k in list(hasil) if _base_produk(k) == "Body Lotion"]
        ss_keys = [k for k in list(hasil) if _base_produk(k) == "Sunscreen"]
        if bl_keys and ss_keys:
            for k in ss_keys:
                del hasil[k]
        boa_keys = [k for k in list(hasil) if k == "BOA Produk" or k.startswith("BOA Produk ")]
        for k in boa_keys:
            del hasil[k]
    return hasil


# ── Companion products: gratis dari segi harga, tapi tetap dipacking ────────
# Key = SKU exact (uppercase), Value = nama produk
_COMPANION_SKUS: dict[str, str] = {
    "0213AMCR": "Men Care",
    "MENCARE":  "Men Care",
}


def parse_companion_produk(text: str) -> dict:
    """
    Deteksi produk companion (free gift tapi real produk) dari teks resi.
    Return dict {nama_produk: total_qty}, terpisah dari produk utama.
    Tidak mempengaruhi grouping/sortir — hanya untuk rekap packing.
    """
    if not text:
        return {}

    text_clean = re.sub(r"(?:Pengirim|Sender)\s*[:\(][^\n]*\n?", "", text, flags=re.IGNORECASE)
    merch_m = re.search(
        r"Merchant\s+Title\s+SKU\s+Qty(.+)",
        text_clean, re.DOTALL | re.IGNORECASE,
    )
    if not merch_m:
        return {}

    hasil: dict[str, int] = {}
    section = merch_m.group(1)
    comp_upper = {k.upper(): v for k, v in _COMPANION_SKUS.items()}

    for line in section.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.search(r"^(.+?)\s+\b([A-Z0-9]{6,}[-A-Z0-9]*)\s+(\d+)\s*$", line)
        if m:
            sku = m.group(2).strip().upper()
            if sku in comp_upper:
                nama = comp_upper[sku]
                hasil[nama] = hasil.get(nama, 0) + int(m.group(3))

    return hasil


# ══════════════════════════════════════════════════════════════
#  REPAIR PDF
# ══════════════════════════════════════════════════════════════

def _repair_pdf(pdf_bytes: bytes) -> bytes:
    try:
        import pikepdf
        with pikepdf.open(io.BytesIO(pdf_bytes)) as pdf:
            out = io.BytesIO()
            pdf.save(out)
            return out.getvalue()
    except Exception:
        pass
    return pdf_bytes


# ══════════════════════════════════════════════════════════════
#  SORT KEY
# ══════════════════════════════════════════════════════════════

def _sort_key(parse_result: dict) -> tuple:
    if not parse_result:
        return ("ZZZ_EMPTY", 999, "")
    products = list(parse_result.keys())
    if len(products) > 1:
        return ("ZZZ_MULTI", 0, " + ".join(sorted(products)))
    prod = products[0]
    m    = re.search(r"(\d+)\s*PCS", prod, re.IGNORECASE)
    qty  = int(m.group(1)) if m else 1
    base = re.sub(r"\s*\d+\s*PCS\s*$", "", prod, flags=re.IGNORECASE).strip().upper()
    return (base or prod.upper(), qty, prod)


def _is_multi(parse_result: dict) -> bool:
    if not parse_result:
        return False
    bases = set()
    for n in parse_result:
        b = re.sub(r"\s*\d+\s*PCS\s*$", "", n, flags=re.IGNORECASE).strip().upper()
        bases.add(b or n.upper())
    return len(bases) > 1


# ══════════════════════════════════════════════════════════════
#  SLUG HELPERS
# ══════════════════════════════════════════════════════════════

def _camel(text: str) -> str:
    parts = re.split(r"\s+", text.strip())
    return "".join((p[0].upper() + p[1:]) if p else "" for p in parts)


def _produk_slug(produk_name: str) -> str:
    s = produk_name.strip()
    m = re.search(r"^(.+?)\s+(\d+)\s*PCS\s*$", s, re.IGNORECASE)
    if m:
        base, qty = m.group(1).strip(), m.group(2)
        return f"{_camel(base)}_{qty}PCS"
    return _camel(s)


_KURIR_FILENAME_ALIAS = {
    "SPX": "SPX", "J&T": "JNT", "JNT LEX": "JNTLEX", "SiCepat": "CPT",
    "JNE": "JNE", "JNE LEX": "JLEX", "Anteraja": "ANTER", "GTL": "GTL",
    "Ninja Xpress": "NP", "Ninja LEX": "NLEX", "Ninja Pusat": "NP",
    "SAP": "SAP", "SAP LEX": "SAPLEX", "ID Express": "ID",
    "Pos Indonesia": "POS", "Lion Parcel": "LION", "Instan": "INSTAN",
    "Sentral Cargo": "SENTRAL",
}


def _kurir_slug(kurir: str) -> str:
    if kurir in _KURIR_FILENAME_ALIAS:
        return _KURIR_FILENAME_ALIAS[kurir]
    return re.sub(r"[^\w\-]", "", kurir.replace("&", ""))


# ══════════════════════════════════════════════════════════════
#  KLOTERAN ENGINE
# ══════════════════════════════════════════════════════════════

# Nama singkatan produk untuk print sheet kloteran
_PRODUK_SINGKAT = {
    "Body Lotion":          "Body Lotion",
    "Body Lotion Sachet":   "BL Sachet",
    "Body Wash":            "Body Wash",
    "Lip Serum":            "Lip Serum",
    "Eye Cream":            "Eye Cream",
    "Eye Wrinkle":          "Eye Wrinkle",
    "Underarm":             "Underarm",
    "Feminine":             "Feminine",
    "Feminine Spray":       "Fem Spray",
    "Men Care":             "Men Care",
    "Sunscreen":            "Sunscreen",
    "Glow Soap":            "Glow Soap",
    "Brightening Serum":    "Bright Serum",
    "Peeling Serum":        "Peeling",
    "Moisturizer Serum":    "Moisturizer",
    "Acne Serum":           "Acne Serum",
    "Parfume":              "Parfume",
    "Parfume BYC":          "BYC",
    "Collagen Drink":       "Collagen",
    "White Tomato Series":  "WT Series",
    "Toner":                "Toner",
    "Facial Wash":          "Facial Wash",
    "Day Cream":            "Day Cream",
    "Night Gel":            "Night Gel",
    "Cleansing Balm":       "Cleansing Balm",
    "Meili Beauty Cream":   "Meili",
}


def _singkat_produk(nama: str) -> str:
    """Ambil nama singkat produk. Strip PCS suffix dulu, lalu lookup."""
    base = re.sub(r"\s*\d+\s*PCS\s*$", "", nama, flags=re.IGNORECASE).strip()
    short = _PRODUK_SINGKAT.get(base, base)
    # Cek apakah ada qty suffix
    m = re.search(r"(\d+)\s*PCS", nama, re.IGNORECASE)
    if m:
        return f"{short} {m.group(1)}pcs"
    return short


def _kloteran_label(nomor: int, tanggal: datetime) -> str:
    """Ganjil → angka (1,2,3), Genap → huruf (A,B,C)."""
    if tanggal.day % 2 == 1:  # ganjil
        return str(nomor)
    else:  # genap
        # 1→A, 2→B, dst
        letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        return letters[nomor - 1] if nomor <= 26 else str(nomor)


def assign_kloteran(
    groups: dict,
    tanggal: datetime,
    batas_single: int = 50,
    batas_multi: int = 10,
) -> list[dict]:
    """
    Auto-assign kloteran dari groups hasil split_pdf_by_kurir.

    Returns list of kloteran dicts:
    {
        "nomor": int,
        "label": str,
        "kurir": str,
        "is_multi": bool,
        "produk_list": [str],       # list produk singkat
        "produk_display": str,      # untuk tabel UI
        "produk_qty": dict,         # {nama_singkat: total_pcs} untuk kartu cetak
        "jumlah_resi": int,
        "total_pcs": int,           # total semua pcs di kloteran ini
        "group_key": str,
        "page_indices": [int],
    }
    """
    kloterans = []
    counter   = 1

    all_keys  = [k for k in sorted(groups.keys()) if not k.startswith("_")]
    non_multi = [k for k in all_keys if not k.endswith("_Multi") and k != KURIR_BELUM_TERDETEKSI]
    multi     = [k for k in all_keys if k.endswith("_Multi")]
    undetect  = [k for k in all_keys if k == KURIR_BELUM_TERDETEKSI]
    ordered   = non_multi + multi + undetect

    for gkey in ordered:
        d = groups[gkey]
        if not d.get("pages"):
            continue

        kurir    = d.get("kurir") or "—"
        is_multi = gkey.endswith("_Multi") or gkey == KURIR_BELUM_TERDETEKSI
        batas    = batas_multi if is_multi else batas_single

        sorted_pages = sorted(d["pages"], key=lambda x: x[2])
        all_pages    = [(idx, parse_result) for idx, parse_result, _ in sorted_pages]

        # Pecah jadi chunk
        chunks_pages = [all_pages[i:i+batas] for i in range(0, len(all_pages), batas)]

        for chunk in chunks_pages:
            chunk_idx = [idx for idx, _ in chunk]

            # Hitung qty per produk dalam chunk ini
            produk_qty: dict[str, int] = {}
            for _, parse_result in chunk:
                for nama, _ in parse_result.items():
                    short = _singkat_produk(nama)
                    m_pcs = re.search(r"(\d+)\s*PCS", nama, re.IGNORECASE)
                    pcs_per_resi = int(m_pcs.group(1)) if m_pcs else 1
                    produk_qty[short] = produk_qty.get(short, 0) + pcs_per_resi

            # Hitung companion per chunk (proporsional dari total group)
            # companion_per_produk di-store di group, bukan per-page
            # → estimasi: total_companion / total_resi_group * jumlah_resi_chunk
            companion_qty: dict[str, int] = {}
            total_resi_group = d["total_resi"]
            for comp_nama, comp_total in d.get("companion_per_produk", {}).items():
                if total_resi_group > 0:
                    # round untuk chunk terakhir — pakai round agar total tepat
                    estimated = round(comp_total * len(chunk_idx) / total_resi_group)
                    if estimated > 0:
                        companion_qty[comp_nama] = estimated

            produk_list = list(produk_qty.keys())
            total_pcs   = sum(produk_qty.values())

            if is_multi:
                produk_display = " / ".join(produk_list)
            else:
                produk_display = produk_list[0] if produk_list else (d.get("produk") or gkey)

            label = _kloteran_label(counter, tanggal)
            kloterans.append({
                "nomor":          counter,
                "label":          label,
                "kurir":          kurir,
                "is_multi":       is_multi,
                "produk_list":    produk_list,
                "produk_display": produk_display,
                "produk_qty":     produk_qty,
                "companion_qty":  companion_qty,   # {nama: qty} — produk gratis real
                "jumlah_resi":    len(chunk_idx),
                "total_pcs":      total_pcs,
                "group_key":      gkey,
                "page_indices":   chunk_idx,
            })
            counter += 1

    return kloterans


def merge_kloteran(kloterans: list[dict], indices: list[int]) -> list[dict]:
    """
    Gabung kloteran pada posisi `indices` jadi 1 kloteran baru.
    Kloteran hasil gabungan ditempatkan di posisi pertama dari indices.
    """
    if len(indices) < 2:
        return kloterans

    indices_set = set(indices)
    to_merge    = [kloterans[i] for i in sorted(indices)]

    merged_produk_qty:    dict[str, int] = {}
    merged_companion_qty: dict[str, int] = {}
    merged_pages:         list[int]      = []
    merged_resi = 0

    for kl in to_merge:
        merged_pages.extend(kl.get("page_indices", []))
        merged_resi += kl["jumlah_resi"]
        for p, n in kl.get("produk_qty", {}).items():
            merged_produk_qty[p] = merged_produk_qty.get(p, 0) + n
        for c, n in kl.get("companion_qty", {}).items():
            merged_companion_qty[c] = merged_companion_qty.get(c, 0) + n

    label       = "+".join(kl["label"] for kl in to_merge)
    kurir       = to_merge[0]["kurir"]
    produk_list = list(merged_produk_qty.keys())
    is_multi    = (
        len({re.sub(r"\s*\d+\s*pcs\s*$", "", p, flags=re.IGNORECASE).strip()
             for p in merged_produk_qty}) > 1
        or any(kl["is_multi"] for kl in to_merge)
    )
    produk_display = (
        " / ".join(produk_list) if is_multi
        else (produk_list[0] if produk_list else "—")
    )

    merged = {
        "nomor":          to_merge[0]["nomor"],
        "label":          label,
        "kurir":          kurir,
        "is_multi":       is_multi,
        "produk_list":    produk_list,
        "produk_display": produk_display,
        "produk_qty":     merged_produk_qty,
        "companion_qty":  merged_companion_qty,
        "jumlah_resi":    merged_resi,
        "total_pcs":      sum(merged_produk_qty.values()),
        "group_key":      "merged",
        "page_indices":   merged_pages,
    }

    # Susun ulang: merged masuk di posisi pertama, sisanya tetap urut
    result: list[dict] = []
    inserted = False
    for i, kl in enumerate(kloterans):
        if i in indices_set:
            if not inserted:
                result.append(merged)
                inserted = True
        else:
            result.append(kl)
    return result


# ══════════════════════════════════════════════════════════════
#  BUILD KLOTERAN EXCEL
# ══════════════════════════════════════════════════════════════

def build_kloteran_excel(
    kloterans: list[dict],
    tanggal: datetime,
    nama_admin: str,
) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kloteran"

    HEAD_FILL  = PatternFill("solid", fgColor="C49166")
    HEAD_FONT  = Font(bold=True, color="15110D", size=11)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    MULTI_FILL = PatternFill("solid", fgColor="2E251E")
    MULTI_FONT = Font(color="D4A574", size=10)

    thin = Side(style="thin", color="3A2F26")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header info
    ws.merge_cells("A1:F1")
    ws["A1"] = f"REKAP KLOTERAN — {tanggal.strftime('%d %B %Y')} — Admin: {nama_admin}"
    ws["A1"].font = Font(bold=True, size=13, color="F0E4D2")
    ws["A1"].fill = PatternFill("solid", fgColor="15110D")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Total Kloteran: {len(kloterans)}"
    ws["A2"].font = Font(size=10, color="8A7A66")
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 18

    # Column headers
    headers = ["Kloteran", "Kurir", "Produk", "Jumlah Resi", "Tipe", "Catatan"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[3].height = 22

    # Data rows
    for i, kl in enumerate(kloterans, start=4):
        is_m  = kl["is_multi"]

        # Bangun catatan: tampilkan detail produk + companion
        pq    = kl.get("produk_qty", {})
        cq    = kl.get("companion_qty", {})
        detail_parts = [f"{p}: {n} pcs" for p, n in pq.items()]
        comp_parts   = [f"{p}: {n} pcs (gratis)" for p, n in cq.items()]
        catatan = "  |  ".join(detail_parts + comp_parts)

        produk_cell = kl["produk_display"]

        row_data = [
            kl["label"],
            kl["kurir"],
            produk_cell,
            kl["jumlah_resi"],
            "Multi" if is_m else "Single",
            catatan,
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            if is_m:
                c.font = MULTI_FONT
                c.fill = MULTI_FILL
            c.alignment = CENTER if col != 3 else LEFT
            c.border = BORDER
        ws.row_dimensions[i].height = 32 if is_m else 22

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
#  BUILD KLOTERAN PDF (Print Sheet)
# ══════════════════════════════════════════════════════════════

def _kartu_html(kl: dict, tgl_str: str, nama_toko: str = "BOA PST") -> str:
    """Render 1 kartu kloteran sebagai HTML — siap digunting."""
    is_m     = kl["is_multi"]
    label    = kl["label"]
    kurir    = kl["kurir"]
    jml_resi = kl["jumlah_resi"]
    produk_qty: dict   = kl.get("produk_qty", {})
    companion_qty: dict = kl.get("companion_qty", {})

    badge = (
        '<span class="badge-multi">MULTI</span>' if is_m
        else '<span class="badge-single">SINGLE</span>'
    )

    # Baris produk utama
    prod_rows = ""
    for p_short, total_p in produk_qty.items():
        prod_rows += (
            f'<div class="pr">'
            f'<span class="pr-name">{p_short}</span>'
            f'<span class="pr-qty">{total_p} pcs</span>'
            f'</div>'
        )
    # Companion (gratis tapi real)
    for c_nama, c_total in companion_qty.items():
        prod_rows += (
            f'<div class="pr">'
            f'<span class="pr-name pr-free">{c_nama}*</span>'
            f'<span class="pr-qty pr-free">{c_total} pcs</span>'
            f'</div>'
        )
    if not prod_rows:
        prod_rows = '<div class="pr"><span class="pr-name">—</span></div>'

    return f"""
<div class="kartu">
  <div class="hdr">
    <div class="hdr-left">
      <span class="brand">{nama_toko}</span>
      <span class="knum">{label}</span>
    </div>
    <span class="tgl">{tgl_str}</span>
    <div class="platform"><span>Shopee</span> / <span>TikTok</span></div>
  </div>
  <div class="core">
    <div class="left-big">
      <span class="jr-lbl">RESI</span>
      <span class="jr-num">{jml_resi}</span>
      <div class="divider-hz"></div>
      <span class="ex-code">{kurir}</span>
      <span class="ex-lbl">KURIR</span>
    </div>
    <div class="produk-list">{prod_rows}</div>
  </div>
  <div class="footer">
    <span class="paraf">Paraf <span class="paraf-line"></span></span>
    {badge}
  </div>
</div>"""


def build_kloteran_print_sheet(
    kloterans: list[dict],
    tanggal: datetime,
    nama_admin: str,
    nama_toko: str = "BOA PST",
) -> tuple:
    """
    Build print sheet grid kartu A4 — 2 kolom, semua kloteran 1 halaman.
    Admin tinggal gunting per kartu, tempel ke tumpukan resi.
    Returns (bytes, "pdf"|"html")
    """
    tgl_str  = tanggal.strftime("%d/%m/%Y")
    kartu_html_list = [_kartu_html(kl, tgl_str, nama_toko) for kl in kloterans]
    kartu_joined    = "\n".join(kartu_html_list)

    total_resi   = sum(k["jumlah_resi"] for k in kloterans)
    total_single = sum(1 for k in kloterans if not k["is_multi"])
    total_multi  = sum(1 for k in kloterans if k["is_multi"])

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<style>
  @page {{ size: A4 portrait; margin: 8mm 8mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    background: #fff;
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 10px;
    color: #15110D;
  }}

  /* ── Page header ── */
  .page-header {{
    display: flex; justify-content: space-between; align-items: center;
    border-bottom: 2px solid #C49166; padding-bottom: 5px; margin-bottom: 8px;
  }}
  .page-header h1 {{ font-size: 13px; font-weight: 700; color: #15110D; }}
  .page-header .meta {{ font-size: 9px; color: #6B5D4D; text-align: right; line-height: 1.6; }}
  .summary-chips {{ display: flex; gap: 6px; margin-bottom: 8px; flex-wrap: wrap; }}
  .chip {{
    background: #F5EFE6; border: 0.5px solid #C49166;
    border-radius: 999px; padding: 2px 9px;
    font-size: 9px; color: #15110D;
  }}
  .chip b {{ color: #A8744E; }}

  /* ── Grid kartu ── */
  .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 6px; }}

  /* ── Kartu ── */
  .kartu {{
    border: 1.5px solid #C49166;
    border-radius: 8px;
    overflow: hidden;
    background: #fff;
    break-inside: avoid;
    page-break-inside: avoid;
  }}

  /* Header baris */
  .hdr {{
    display: flex; justify-content: space-between; align-items: center;
    padding: 4px 7px; gap: 5px;
    border-bottom: 0.5px solid #D4B896;
  }}
  .hdr-left {{ display: flex; align-items: center; gap: 4px; flex-shrink: 0; }}
  .brand {{ font-size: 8px; font-weight: 700; color: #A8744E; letter-spacing: .3px; }}
  .knum {{
    font-size: 11px; font-weight: 700; color: #15110D;
    background: #F5EFE6; border: 0.5px solid #D4B896;
    border-radius: 3px; padding: 0 5px; line-height: 1.5;
  }}
  .tgl {{ font-size: 8px; color: #6B5D4D; white-space: nowrap; }}
  .platform {{ font-size: 8px; color: #B8A78D; white-space: nowrap; }}
  .platform span {{
    padding: 0 4px; border: 0.5px solid #D4B896; border-radius: 3px;
  }}

  /* Core: kiri besar + kanan produk */
  .core {{ display: flex; border-bottom: 0.5px solid #D4B896; }}

  .left-big {{
    width: 58px; min-width: 58px;
    display: flex; flex-direction: column;
    align-items: center; justify-content: center;
    padding: 8px 0;
    border-right: 0.5px solid #D4B896;
    gap: 0;
  }}
  .jr-lbl {{ font-size: 7px; color: #B8A78D; letter-spacing: .4px; margin-bottom: 1px; }}
  .jr-num {{ font-size: 36px; font-weight: 700; color: #15110D; line-height: 1; }}
  .divider-hz {{ width: 38px; height: 1px; background: #C49166; margin: 4px 0; }}
  .ex-code {{ font-size: 18px; font-weight: 700; color: #15110D; line-height: 1; }}
  .ex-lbl {{ font-size: 7px; color: #B8A78D; letter-spacing: .4px; margin-top: 2px; }}

  .produk-list {{ flex: 1; padding: 5px 7px; display: flex; flex-direction: column; gap: 3px; }}
  .pr {{
    display: flex; justify-content: space-between; align-items: baseline;
    border-bottom: 0.5px dotted #EDE0D0; padding-bottom: 2px;
  }}
  .pr:last-child {{ border-bottom: none; padding-bottom: 0; }}
  .pr-name {{ font-size: 9px; color: #15110D; max-width: 85px; line-height: 1.3; }}
  .pr-name.pr-free {{ color: #8A7A66; font-style: italic; }}
  .pr-qty {{ font-size: 10px; font-weight: 700; color: #15110D; white-space: nowrap; }}
  .pr-qty.pr-free {{ color: #8A7A66; font-weight: 400; }}

  /* Footer */
  .footer {{
    display: flex; justify-content: space-between; align-items: center;
    padding: 4px 7px;
  }}
  .paraf {{ font-size: 8px; color: #B8A78D; }}
  .paraf-line {{
    display: inline-block; width: 60px; height: 0.5px;
    background: #C49166; vertical-align: middle; margin-left: 3px;
  }}
  .badge-multi {{
    font-size: 7px; font-weight: 700;
    padding: 1px 5px; border-radius: 3px;
    background: #C49166; color: #15110D;
  }}
  .badge-single {{
    font-size: 7px; font-weight: 700;
    padding: 1px 5px; border-radius: 3px;
    background: #F5EFE6; color: #A8744E;
    border: 0.5px solid #C49166;
  }}

  /* ── Footer halaman ── */
  .page-footer {{
    margin-top: 8px; font-size: 8px; color: #B8A78D;
    border-top: 0.5px solid #EDE0D0; padding-top: 4px;
    display: flex; justify-content: space-between;
  }}
</style>
</head>
<body>

<div class="page-header">
  <div>
    <h1>Rekap Kloteran — {nama_toko}</h1>
    <div style="font-size:9px;color:#8A7A66;">Admin: {nama_admin} &nbsp;·&nbsp; {tanggal.strftime('%d %B %Y')}</div>
  </div>
  <div class="meta">
    BOA Sortir Tools<br/>
    Print: {datetime.now().strftime('%H:%M')} WIB
  </div>
</div>

<div class="summary-chips">
  <div class="chip">Total Kloteran <b>{len(kloterans)}</b></div>
  <div class="chip">Total Resi <b>{total_resi}</b></div>
  <div class="chip">Single <b>{total_single}</b></div>
  <div class="chip">Multi <b>{total_multi}</b></div>
</div>

<div class="grid">
{kartu_joined}
</div>

<div class="page-footer">
  <span>BOA Sortir Tools v1.2 — Crafted by Mashori</span>
  <span>Dokumen internal — potong per kotak, tempel ke tumpukan resi</span>
</div>

</body>
</html>"""

    try:
        from weasyprint import HTML as WP_HTML
        pdf_bytes = WP_HTML(string=html).write_pdf()
        return pdf_bytes, "pdf"
    except Exception:
        return html.encode("utf-8"), "html"


# Alias untuk backward compat dengan tombol lama
def build_kloteran_pdf(kloterans, tanggal, nama_admin):
    return build_kloteran_print_sheet(kloterans, tanggal, nama_admin)


# ══════════════════════════════════════════════════════════════
#  BUILD ZIP PDF RESI PER KLOTERAN
# ══════════════════════════════════════════════════════════════

def build_kloteran_zip_pdf(
    pdf_bytes: bytes,
    kloterans: list[dict],
) -> bytes:
    """
    Buat ZIP berisi PDF resi asli yang dipotong per kloteran.
    Nama file: Kloteran_{label}_{kurir}_{produk_slug}.pdf
    """
    from pypdf import PdfReader, PdfWriter
    pdf_bytes = _repair_pdf(pdf_bytes)
    reader    = PdfReader(io.BytesIO(pdf_bytes))

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for kl in kloterans:
            writer = PdfWriter()
            for idx in kl["page_indices"]:
                if idx < len(reader.pages):
                    writer.add_page(reader.pages[idx])

            pdf_buf = io.BytesIO()
            writer.write(pdf_buf)

            # Nama file
            kurir_slug  = re.sub(r"[^\w]", "", kl["kurir"].replace("&", ""))
            produk_slug = re.sub(r"[^\w]", "_", kl["produk_display"])[:30]
            fname       = f"Kloteran_{kl['label']}_{kurir_slug}_{produk_slug}.pdf"
            zf.writestr(fname, pdf_buf.getvalue())

    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
#  CORE: SPLIT PDF BY KURIR
# ══════════════════════════════════════════════════════════════

def split_pdf_by_kurir(
    pdf_bytes: bytes,
    progress_cb=None,
    use_ocr: bool = True,
    use_groq: bool = False,
    groq_api_key: str | None = None,
) -> dict:
    from pypdf import PdfReader

    pdf_bytes   = _repair_pdf(pdf_bytes)
    reader      = PdfReader(io.BytesIO(pdf_bytes))
    total_pages = len(reader.pages)

    method_counts = defaultdict(int)
    pages_info    = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf_pl:
        for i in range(total_pages):
            if progress_cb:
                progress_cb(i, total_pages, f"Memproses halaman {i+1}/{total_pages}…")

            page = pdf_pl.pages[i]
            text = page.extract_text() or ""

            kurir  = detect_kurir(text)
            method = "text" if kurir else "none"

            if not kurir and use_ocr:
                png = _render_page_png(pdf_bytes, i)
                if png:
                    ocr_text = _ocr_logo(png)
                    if ocr_text:
                        kurir = detect_kurir(ocr_text)
                        if kurir:
                            method = "ocr"

            if not kurir and use_groq:
                png = _render_page_png(pdf_bytes, i, dpi=150)
                if png:
                    g_kurir = _groq_detect_kurir(png, api_key=groq_api_key)
                    if g_kurir:
                        kurir  = g_kurir
                        method = "groq"

            method_counts[method] += 1

            parse_result     = parse_produk(text)
            companion_result = parse_companion_produk(text)

            # Kalau produk utama IS companion (e.g. resi Men Care murni),
            # jangan double-count di companion_result
            for main_key in list(parse_result.keys()):
                base = re.sub(r"\s*\d+\s*PCS\s*$", "", main_key, flags=re.IGNORECASE).strip()
                companion_result.pop(base, None)

            pages_info.append({
                "idx":              i,
                "kurir":            kurir,
                "parse_result":     parse_result,
                "companion_result": companion_result,
                "method":           method,
                "is_multi_sku":     _is_multi(parse_result),
            })

    if progress_cb:
        progress_cb(total_pages, total_pages, "Mengelompokkan resi…")

    sku_count = defaultdict(int)
    for p in pages_info:
        if p["kurir"] and not p["is_multi_sku"] and p["parse_result"]:
            sku_name = next(iter(p["parse_result"].keys()))
            sku_count[(p["kurir"], sku_name)] += 1

    groups = defaultdict(lambda: {
        "pages": [],
        "total_resi": 0,
        "total_per_produk": defaultdict(int),
        "companion_per_produk": defaultdict(int),   # produk gratis tapi real
        "kurir": None,
        "produk": None,
    })

    for p in pages_info:
        kurir        = p["kurir"]
        parse_result = p["parse_result"]
        idx          = p["idx"]
        sk           = _sort_key(parse_result)

        if not kurir:
            target = KURIR_BELUM_TERDETEKSI
            grp_kurir, grp_produk = None, None
        elif p["is_multi_sku"] or not parse_result:
            target = f"{_kurir_slug(kurir)}_Multi"
            grp_kurir, grp_produk = kurir, "Mixed (orphan + multi-SKU)"
        else:
            sku_name = next(iter(parse_result.keys()))
            count    = sku_count[(kurir, sku_name)]
            if count >= 2:
                target = f"{_kurir_slug(kurir)}_{_produk_slug(sku_name)}"
                grp_kurir, grp_produk = kurir, sku_name
            else:
                target = f"{_kurir_slug(kurir)}_Multi"
                grp_kurir, grp_produk = kurir, "Mixed (orphan + multi-SKU)"

        g = groups[target]
        g["pages"].append((idx, parse_result, sk))
        g["total_resi"] += 1
        g["kurir"]  = grp_kurir
        g["produk"] = grp_produk

        for nama, qty in parse_result.items():
            m_pcs  = re.search(r"(\d+)\s*PCS", nama, re.IGNORECASE)
            batang = int(m_pcs.group(1)) if m_pcs else 1
            g["total_per_produk"][nama] += qty
            prev = g["total_per_produk"].get(f"_BATANG_{nama}", 0)
            g["total_per_produk"][f"_BATANG_{nama}"] = prev + batang * qty

        # Akumulasi companion (gratis tapi real product)
        for nama, qty in p.get("companion_result", {}).items():
            g["companion_per_produk"][nama] += qty

    if progress_cb:
        progress_cb(total_pages, total_pages, "Klasifikasi selesai.")

    result = dict(groups)
    result["_meta"] = {
        "method_counts": dict(method_counts),
        "total_pages":   total_pages,
    }
    return result


# ══════════════════════════════════════════════════════════════
#  BUILD OUTPUT
# ══════════════════════════════════════════════════════════════

def build_output_pdfs(pdf_bytes: bytes, groups: dict) -> dict:
    from pypdf import PdfReader, PdfWriter
    pdf_bytes = _repair_pdf(pdf_bytes)
    reader    = PdfReader(io.BytesIO(pdf_bytes))
    out       = {}
    for name, data in groups.items():
        if name.startswith("_") or not data["pages"]:
            continue
        sorted_pages = sorted(data["pages"], key=lambda x: x[2])
        writer = PdfWriter()
        for idx, _, _ in sorted_pages:
            writer.add_page(reader.pages[idx])
        buf = io.BytesIO()
        writer.write(buf)
        out[name] = buf.getvalue()
    return out


def build_excel_rekap(groups: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    HEAD_FILL = PatternFill("solid", fgColor="C49166")
    HEAD_FONT = Font(bold=True, color="15110D", size=11)
    CENTER    = Alignment(horizontal="center", vertical="center")

    ws1 = wb.active
    ws1.title = "Ringkasan"
    ws1.append(["File PDF", "Kurir", "Produk / Kategori", "Total Resi", "Total Batang"])
    for c in ws1[1]:
        c.font = HEAD_FONT; c.fill = HEAD_FILL; c.alignment = CENTER

    sorted_keys = sorted(groups.keys())
    for name in sorted_keys:
        if name.startswith("_"):
            continue
        d = groups[name]
        if not d.get("pages"):
            continue
        kurir   = d.get("kurir") or "—"
        produk  = d.get("produk") or "Mixed"
        total_b = sum(v for k, v in d["total_per_produk"].items() if k.startswith("_BATANG_"))
        ws1.append([f"{name}.pdf", kurir, produk, d["total_resi"], total_b])

    ws2 = wb.create_sheet("Detail")
    ws2.append(["File PDF", "Kurir", "Produk", "Jumlah Resi", "Total Batang"])
    for c in ws2[1]:
        c.font = HEAD_FONT; c.fill = HEAD_FILL; c.alignment = CENTER

    for name in sorted_keys:
        if name.startswith("_"):
            continue
        d = groups[name]
        if not d.get("pages"):
            continue
        kurir = d.get("kurir") or "—"
        for prod, n_resi in sorted(d["total_per_produk"].items()):
            if prod.startswith("_BATANG_"):
                continue
            n_batang = d["total_per_produk"].get(f"_BATANG_{prod}", n_resi)
            ws2.append([f"{name}.pdf", kurir, prod, n_resi, n_batang])

    for ws in (ws1, ws2):
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_zip(output_pdfs: dict, excel_bytes: bytes | None = None) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, pdf_b in output_pdfs.items():
            safe = re.sub(r"[^\w\-]", "_", name)
            zf.writestr(f"{safe}.pdf", pdf_b)
        if excel_bytes:
            zf.writestr("Rekap.xlsx", excel_bytes)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
#  SIDEBAR — SETTINGS KLOTERAN
# ══════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("""
    <div style="font-size:13px;font-weight:700;color:#C49166;margin-bottom:8px;">
        ⚙️ Settings Kloteran
    </div>
    """, unsafe_allow_html=True)

    tgl_default = (datetime.utcnow() + timedelta(hours=7)).date()
    tgl_kloteran = st.date_input("Tanggal Sortir", value=tgl_default, key="tgl_kloteran")
    nama_admin   = st.text_input("Nama Admin", value="Admin", key="nama_admin",
                                  placeholder="Nama admin / operator")
    batas_single = st.number_input("Batas resi Single/kloteran", min_value=1, max_value=200,
                                    value=50, step=5, key="batas_single",
                                    help="Jumlah maksimal resi 1 produk per kloteran")
    batas_multi  = st.number_input("Batas resi Multi/kloteran", min_value=1, max_value=50,
                                    value=10, step=1, key="batas_multi",
                                    help="Jumlah maksimal resi campur per kloteran (lebih kecil, lebih teliti)")

    st.divider()
    st.markdown("""
    <div style="font-size:11px;color:#6B5D4D;line-height:1.6;">
    <b style="color:#C9B89F;">Penomoran:</b><br/>
    Tanggal ganjil → 1, 2, 3 …<br/>
    Tanggal genap → A, B, C …
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  MAIN UI  (sudah login)
# ══════════════════════════════════════════════════════════════

col_h1, col_h2 = st.columns([5, 1])
with col_h1:
    st.markdown("""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:.25rem;">
        <div style="width:42px;height:42px;border-radius:10px;
                    background:linear-gradient(135deg,#C49166 0%,#A8744E 100%);
                    display:flex;align-items:center;justify-content:center;
                    color:#15110D;font-weight:700;font-size:20px;
                    box-shadow:0 2px 8px rgba(196,145,102,0.3);">S</div>
        <div>
            <div style="font-size:20px;font-weight:700;line-height:1.2;color:#F0E4D2;">BOA Sortir Tools</div>
            <div style="font-size:12px;color:#8A7A66;">Pemisah resi otomatis per ekspedisi + Kloteran</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
with col_h2:
    if st.button("Keluar", key="btn_logout"):
        _logout()
        st.rerun()

st.divider()

# ── Tabs: Sortir & Kloteran ──
tab_sortir, tab_kloteran = st.tabs(["📤 Sortir Resi", "📋 Kloteran"])

# ══════════════════════════════════════════════════════════════
#  TAB 1: SORTIR RESI
# ══════════════════════════════════════════════════════════════

with tab_sortir:
    st.markdown("#### 📤 Upload PDF Resi")
    st.caption(
        "Upload **1 file PDF** berisi label resi campuran dari berbagai ekspedisi. "
        "Tools akan memisahnya menjadi PDF per kurir, lalu di-bundle ke 1 file ZIP."
    )

    uploaded = st.file_uploader(
        "Pilih file PDF",
        type=["pdf"],
        key="sortir_upload",
        label_visibility="collapsed",
    )

    _GROQ_KEY_AVAILABLE = bool(_get_groq_api_key())

    col_o1, col_o2, col_o3 = st.columns(3)
    with col_o1:
        gen_excel = st.checkbox("📊 Rekap Excel", value=True, key="opt_excel",
                                help="Sertakan file Rekap.xlsx di dalam ZIP")
    with col_o2:
        use_ocr = st.checkbox("🔍 OCR Tesseract", value=True, key="opt_ocr",
                               help="Layer 2: OCR pojok kanan-atas (Tesseract).")
    with col_o3:
        use_groq = st.checkbox(
            "🤖 Groq Vision",
            value=_GROQ_KEY_AVAILABLE,
            disabled=not _GROQ_KEY_AVAILABLE,
            key="opt_groq",
            help=(
                "Layer 3: Groq Vision (Llama 4 Scout)."
                if _GROQ_KEY_AVAILABLE else
                "Disabled — GROQ_API_KEY tidak ditemukan."
            ),
        )

    if uploaded is None:
        st.info("⬆️ Upload PDF resi dulu untuk memulai.")
    elif st.button("🚀 Mulai Sortir", type="primary", use_container_width=True, key="btn_run"):
        pdf_bytes = uploaded.read()

        progress_bar = st.progress(0.0)
        status_text  = st.empty()

        def _cb(current, total, msg):
            progress_bar.progress(min(current / max(total, 1), 1.0))
            status_text.text(msg)

        with st.spinner("Mengklasifikasi resi per kurir…"):
            try:
                groups = split_pdf_by_kurir(
                    pdf_bytes, progress_cb=_cb,
                    use_ocr=use_ocr, use_groq=use_groq,
                )
            except Exception as e:
                st.error(f"❌ Gagal memproses PDF: {e}")
                st.stop()

        progress_bar.progress(1.0)
        status_text.empty()

        meta = groups.pop("_meta", {})

        if not groups:
            st.error("❌ Tidak ada halaman yang berhasil diproses.")
            st.stop()

        # Simpan ke session state supaya bisa dipakai di tab Kloteran
        st.session_state["groups"]    = groups
        st.session_state["pdf_bytes"] = pdf_bytes

        mc     = meta.get("method_counts", {})
        labels = {"text": "📄 Text", "ocr": "🔍 OCR", "groq": "🤖 Groq", "none": "❓ Tidak terdeteksi"}
        parts  = [f"{labels.get(m, m)}: **{n}**" for m, n in mc.items() if n > 0]
        if parts:
            st.caption("Metode deteksi: " + " · ".join(parts))

        st.markdown("#### 📊 Ringkasan Hasil Sortir")
        rows = []
        for name in sorted(groups.keys()):
            d       = groups[name]
            kurir   = d.get("kurir") or "—"
            produk  = d.get("produk") or "Mixed"
            total_b = sum(v for k, v in d["total_per_produk"].items() if k.startswith("_BATANG_"))
            rows.append({
                "File PDF": f"{name}.pdf",
                "Kurir": kurir,
                "Produk / Kategori": produk,
                "Total Resi": d["total_resi"],
                "Total Batang": total_b,
            })
        st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)

        with st.spinner("Menyusun PDF tersortir…"):
            output_pdfs = build_output_pdfs(pdf_bytes, groups)

        excel_bytes = None
        if gen_excel:
            with st.spinner("Membuat rekap Excel…"):
                try:
                    excel_bytes = build_excel_rekap(groups)
                except Exception:
                    excel_bytes = None
                    st.warning("⚠️ openpyxl tidak terpasang — ZIP dibuat tanpa Excel.")

        zip_bytes = build_zip(output_pdfs, excel_bytes)
        timestamp = (datetime.utcnow() + timedelta(hours=7)).strftime("%Y%m%d_%H%M")

        st.success(
            f"✅ Selesai! **{len(output_pdfs)} PDF** kurir tersortir"
            + (" + rekap Excel" if excel_bytes else "")
            + " — siap download."
        )

        st.download_button(
            label=f"⬇️ Download sortir_resi_{timestamp}.zip",
            data=zip_bytes,
            file_name=f"sortir_resi_{timestamp}.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )

        with st.expander("📋 Detail per kurir / kategori"):
            for name in sorted(groups.keys()):
                d = groups[name]
                st.markdown(f"**{name}** — {d['total_resi']} resi")
                detail = [
                    {
                        "Produk": prod,
                        "Jumlah Resi": n,
                        "Total Batang": d["total_per_produk"].get(f"_BATANG_{prod}", n),
                    }
                    for prod, n in sorted(d["total_per_produk"].items())
                    if not prod.startswith("_BATANG_")
                ]
                if detail:
                    st.dataframe(pd.DataFrame(detail), hide_index=True, use_container_width=True)
                st.markdown("---")

        st.info("💡 Buka tab **📋 Kloteran** untuk generate rekap kloteran dari hasil sortir ini.")


# ══════════════════════════════════════════════════════════════
#  TAB 2: KLOTERAN
# ══════════════════════════════════════════════════════════════

with tab_kloteran:
    st.markdown("#### 📋 Rekap Kloteran")

    groups    = st.session_state.get("groups", None)
    pdf_bytes = st.session_state.get("pdf_bytes", None)

    if groups is None:
        st.info("⬅️ Jalankan sortir dulu di tab **📤 Sortir Resi**, lalu kembali ke sini.")
        st.stop()

    tgl_dt      = datetime.combine(tgl_kloteran, datetime.min.time())
    tgl_display = tgl_dt.strftime("%d %B %Y")
    hari_parity = "Ganjil → 1,2,3" if tgl_dt.day % 2 == 1 else "Genap → A,B,C"

    st.caption(
        f"Tanggal: **{tgl_display}** · Admin: **{nama_admin}** · "
        f"Penomoran: **{hari_parity}** · "
        f"Batas single: **{batas_single} resi** · Batas multi: **{batas_multi} resi**"
    )

    # Hitung kloteran awal (fresh dari assign_kloteran)
    kloterans_awal = assign_kloteran(
        groups, tanggal=tgl_dt,
        batas_single=batas_single, batas_multi=batas_multi,
    )

    if not kloterans_awal:
        st.warning("Tidak ada data kloteran. Pastikan sortir sudah dijalankan.")
        st.stop()

    # ── Session state: working kloterans (bisa diedit/digabung) ──
    # Reset kalau upload baru atau parameter berubah
    _kl_params = (id(groups), batas_single, batas_multi, tgl_dt.date())
    if st.session_state.get("_kl_params") != _kl_params:
        st.session_state["kloterans_working"] = kloterans_awal
        st.session_state["_kl_params"]        = _kl_params

    kloterans = st.session_state["kloterans_working"]

    # ── Summary chips ──
    total_resi_kl = sum(k["jumlah_resi"] for k in kloterans)
    col_s1, col_s2, col_s3, col_s4 = st.columns(4)
    col_s1.metric("Total Kloteran", len(kloterans))
    col_s2.metric("Total Resi",     total_resi_kl)
    col_s3.metric("Single", sum(1 for k in kloterans if not k["is_multi"]))
    col_s4.metric("Multi",  sum(1 for k in kloterans if k["is_multi"]))

    # ── Tabel dengan checkbox Pilih ──
    df_kl = pd.DataFrame([{
        "Pilih":     False,
        "Kloteran":  kl["label"],
        "Kurir":     kl["kurir"],
        "Produk":    kl["produk_display"],
        "Jml Resi":  kl["jumlah_resi"],
        "Total PCS": kl.get("total_pcs", 0),
        "Tipe":      "Multi" if kl["is_multi"] else "Single",
    } for kl in kloterans])

    edited_df = st.data_editor(
        df_kl,
        column_config={
            "Pilih":     st.column_config.CheckboxColumn("☑", width="small"),
            "Kloteran":  st.column_config.TextColumn("Kloteran", width="small"),
            "Tipe":      st.column_config.TextColumn("Tipe", width="small"),
        },
        disabled=["Kloteran", "Kurir", "Produk", "Jml Resi", "Total PCS", "Tipe"],
        hide_index=True,
        use_container_width=True,
        key="kl_table_editor",
    )

    # ── Gabung UI ──
    selected_indices = [int(i) for i, row in edited_df.iterrows() if row["Pilih"]]
    n_sel = len(selected_indices)

    col_info, col_reset = st.columns([4, 1])
    with col_info:
        if n_sel >= 2:
            sel_labels = " + ".join(kloterans[i]["label"] for i in selected_indices)
            st.caption(f"✅ **{n_sel} kloteran dipilih:** {sel_labels} — siap digabung")
        elif n_sel == 1:
            st.caption("☝️ Pilih minimal 2 kloteran untuk digabung.")
        else:
            st.caption("☑️ Centang 2 atau lebih baris di tabel untuk menggabungkan kloteran.")
    with col_reset:
        if st.button("🔁 Reset", use_container_width=True, key="btn_kl_reset",
                     help="Kembalikan ke kloteran awal sebelum ada penggabungan"):
            st.session_state["kloterans_working"] = kloterans_awal
            st.rerun()

    if n_sel >= 2:
        sel_labels = " + ".join(kloterans[i]["label"] for i in selected_indices)
        if st.button(
            f"🔗 Gabung: {sel_labels}",
            type="primary",
            use_container_width=True,
            key="btn_gabung",
        ):
            st.session_state["kloterans_working"] = merge_kloteran(kloterans, selected_indices)
            st.rerun()

    st.divider()
    st.markdown("##### 📥 Download Output")

    # ── Row 1: Excel + Print Sheet ──
    col_e, col_p = st.columns(2)

    with col_e:
        if st.button("📊 Excel Kloteran", use_container_width=True, key="btn_kl_excel"):
            with st.spinner("Membuat Excel kloteran…"):
                try:
                    xl_bytes = build_kloteran_excel(kloterans, tgl_dt, nama_admin)
                    fname    = f"kloteran_{tgl_dt.strftime('%Y%m%d')}_{nama_admin}.xlsx"
                    st.download_button(
                        label=f"⬇️ Download {fname}",
                        data=xl_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key="dl_kl_excel",
                    )
                except Exception as ex:
                    st.error(f"❌ Gagal buat Excel: {ex}")

    with col_p:
        if st.button("🖨️ Print Sheet (Kartu Gunting)", use_container_width=True, key="btn_kl_pdf"):
            with st.spinner("Membuat print sheet kartu…"):
                try:
                    out_bytes, out_type = build_kloteran_print_sheet(
                        kloterans, tgl_dt, nama_admin
                    )
                    if out_type == "pdf":
                        fname     = f"printsheet_{tgl_dt.strftime('%Y%m%d')}_{nama_admin}.pdf"
                        mime      = "application/pdf"
                        btn_label = f"⬇️ Download {fname}"
                    else:
                        fname     = f"printsheet_{tgl_dt.strftime('%Y%m%d')}_{nama_admin}.html"
                        mime      = "text/html"
                        btn_label = "⬇️ Download Print Sheet (HTML)"
                    st.download_button(
                        label=btn_label, data=out_bytes, file_name=fname, mime=mime,
                        type="primary", use_container_width=True, key="dl_kl_pdf",
                    )
                    if out_type == "html":
                        st.caption(
                            "ℹ️ WeasyPrint belum terpasang — buka HTML di browser "
                            "lalu Ctrl+P → Save as PDF → gunting per kartu."
                        )
                except Exception as ex:
                    st.error(f"❌ Gagal buat print sheet: {ex}")

    # ── Row 2: PDF resi per kloteran ──
    st.markdown("")
    if st.button(
        "📦 PDF Resi per Kloteran (ZIP)",
        use_container_width=True,
        key="btn_kl_zip",
        help="Buat ZIP berisi PDF resi asli yang sudah dipotong sesuai kloteran.",
    ):
        if pdf_bytes is None:
            st.error("❌ PDF resi asli tidak ditemukan. Upload ulang di tab Sortir Resi.")
        else:
            with st.spinner("Memotong PDF resi per kloteran…"):
                try:
                    zip_kl = build_kloteran_zip_pdf(pdf_bytes, kloterans)
                    fname  = f"resi_kloteran_{tgl_dt.strftime('%Y%m%d')}_{nama_admin}.zip"
                    st.download_button(
                        label=f"⬇️ Download {fname}",
                        data=zip_kl, file_name=fname, mime="application/zip",
                        type="primary", use_container_width=True, key="dl_kl_zip",
                    )
                    st.caption(
                        f"ZIP berisi {len(kloterans)} file PDF — "
                        "nama file: Kloteran_{label}_{kurir}_{produk}.pdf"
                    )
                except Exception as ex:
                    st.error(f"❌ Gagal buat ZIP: {ex}")

    st.divider()

    # ── Detail expand ──
    with st.expander("🔍 Detail per kloteran"):
        for kl in kloterans:
            is_m   = kl["is_multi"]
            tipe   = "Multi" if is_m else "Single"
            pq     = kl.get("produk_qty", {})
            cq     = kl.get("companion_qty", {})
            pq_str = "  ·  ".join(f"{p}: {n} pcs" for p, n in pq.items())
            cq_str = ("  +  " + "  ·  ".join(f"{p}: {n} pcs (gratis)" for p, n in cq.items())) if cq else ""
            st.markdown(
                f"**Kloteran {kl['label']}** — {kl['kurir']} — {tipe} — "
                f"{kl['jumlah_resi']} resi\n\n"
                f"> {pq_str if pq_str else kl['produk_display']}{cq_str}"
            )
            st.markdown("---")
